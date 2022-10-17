
from datetime import datetime, timedelta
import os
# from tracemalloc import start
import openpyxl
import copy

from pathlib import Path

from .models import *
from django.db.models import Q

from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.dimensions import DimensionHolder

from report.class_stoat_report import stoat_report

from datetime import datetime
from calendar import monthrange
from openpyxl.styles import colors
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.formula.translate import Translator
import re

BASE_DIR = Path(__file__).resolve().parent.parent


# Функция дублирующая значения списка, использую для дат
def multiply_items(l):
    return [i for t in zip(l, l) for i in t]

#получение уникальных значение списка
def get_unique_numbers(numbers):
    unique = []
    for number in numbers:
        if number in unique:
            continue
        else:
            unique.append(number)
    return unique

def merge_certain_col(sheet, data, n, c, startcell=10): # c — столбец, n — элемент массива
    mergecount=0
    for row in range(1, len(data)):
        if data[row-1][n] == data[row][n]:
            mergecount += 1
        else:
            if mergecount > 0:
                sheet.merge_cells(start_row=startcell, start_column=c, end_row=startcell+mergecount, end_column=c)
                sheet.cell(row=startcell, column=c).alignment = Alignment(vertical='center')
            mergecount = 0
            startcell = row+10
    if mergecount > 0:
        sheet.merge_cells(start_row=startcell, start_column=c, end_row=startcell+mergecount, end_column=c) 
        sheet.cell(row=startcell, column=c).alignment = Alignment(vertical='center')

class report_po0(stoat_report):
    """ Класс экспорта для Журнала движения и сырья """
    name = 'journal' # имя файла
    start_time = '' # дата начала
    end_time = '' # дата окончания
    end_time_clean = '' # дата окончания
    days = ''
    template = 'media/Unload/common_template.xlsx'
    app = 'move_journal'
    
    
    def add_header(self):
        super().add_header()
        # привожу дату в правильную форму
        self.start_time = datetime.strptime(self.start + '-01 00:00:00', '%Y-%m-%d %H:%M:%S')
        
        if self.stop != None:
            self.end_time = datetime.strptime(
                self.stop + '-01 00:00:00', '%Y-%m-%d %H:%M:%S')
            if self.end_time.month != 12:
                self.end_time_clean = datetime(self.end_time.year, self.end_time.month+1, 1)-timedelta(seconds=1)
            else:
                self.end_time_clean = datetime(self.end_time.year+1, 1, 1)-timedelta(seconds=1)
        else:
            if self.start_time.month != 12:
                self.end_time_clean = datetime(self.start_time.year, self.start_time.month+1, 1)-timedelta(seconds=1)
            else:
                self.end_time_clean = datetime(self.start_time.year+1, 1, 1)-timedelta(seconds=1)

        months = list(range(self.start_time.month, self.end_time_clean.month+1))
        
        # дни
        days_p = []
        for month in months:
            days = monthrange(self.start_time.year, month)
            print('month ', month, ' days ', days[1])
            days_p.extend([f'{str(n).rjust(2, "0")}.{str(month).rjust(2, "0")}.{self.start_time.year}' for n in range(1, days[1]+1)])
        
        # print(days_p)
        self.days = days_p


        #н/с, д/с
        shift = ['н/с', 'д/с']*len(days_p)   
        


    def add_table(self):
        super().add_table()
        double_days = multiply_items(self.days)

        print(' очень странно ')
        print('список всех листов на кирилл ', self.wb.sheetnames)
        # Создаем книги по названиям
        menu = Quide.objects.filter(Q(ancestor=None))
        # menu = ['Droblenie', 'Pererabotka', 'HKF', \
        #         'Filtrovannii k-t', 'Flotokonsentrat', \
        #         'Rachet v probe sliv 25', 'Otcev obrazovannii', \
        #         'Rachet v probe HKF', 'Raschet v probe otseva']
        # -----------------
        for i in menu:
            # print(self.encode('utf8'))
            self.ws = self.wb[str(i)]
            if str(i) != 'Расчет в пробе слив 25' and str(i) != 'Расчет в пробе ХКФ' and str(i) != 'Расчет в пробе отсева':
                for j in range(len(self.days)):
                    self.styled_cell(self.ws.cell(j+4, 2).coordinate, f'{self.days[j]}', 'simple_text')
                    print('type - ', type(self.days[j]), 'value - ', self.days[j])
            else:
                
                for j in range(len(double_days)):
                    self.styled_cell(self.ws.cell(j+4, 2).coordinate, f'{double_days[j]}', 'simple_text')
            
            if str(i) != 'Расчет в пробе слив 25' and str(i) != 'Расчет в пробе ХКФ' and str(i) != 'Расчет в пробе отсева':
                for j in range(1, len(self.days)):
                    # растягиваю формулы
                    for n in range(self.ws.max_column-2):
                        self.ws[self.ws.cell(j+4, 3+n).coordinate] = Translator(self.ws.cell(4, 3+n).value, self.ws.cell(
                            4, 3+n).coordinate).translate_formula(self.ws.cell(j+4, 3+n).coordinate)
            else:
                for j in range(1, len(double_days)-2, 2):
                    # растягиваю формулы
                    for n in range(self.ws.max_column-2):
                        self.ws[self.ws.cell(j+5, 3+n).coordinate] = Translator(self.ws.cell(4, 3+n).value, self.ws.cell(
                            4, 3+n).coordinate).translate_formula(self.ws.cell(j+5, 3+n).coordinate)
                        self.ws[self.ws.cell(j+6, 3+n).coordinate] = Translator(self.ws.cell(5, 3+n).value, self.ws.cell(
                            6, 3+n).coordinate).translate_formula(self.ws.cell(j+5, 3+n).coordinate)
                    
        # --------------------------




        # self.styled_cell(self.ws.cell(4, 1).coordinate, '=ВЕБСЛУЖБА(""&$C$1&"?date="&A7&"&name=Дробление/ВМТ/с.н.м.")', 'simple_text')
        # =ВЕБСЛУЖБА("" &$C$1 & "?date=" & A7 & "&name=Дробление/ВМТ/с.н.м.")
        # print('self.wb.active', self.wb.active)
        #     coord = 0
        #     day = datetime.strptime(i, "%d.%m.%Y")
        #     print('date', day, ' — ', type(day))
        #     type_shift = ['н/с', 'д/с']
        
                            
        #         self.styled_cell(self.ws.cell(7, 1).coordinate, f'Исполнитель: {responsible}', 'simple_text_bold3') 
        #         self.ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=6)  
                
        #         self.styled_cell(self.ws.cell(7, 7).coordinate, f'Время начала сушки: {start_time_dry}', 'simple_text_bold3') 
        #         self.ws.merge_cells(start_row=7, start_column=7, end_row=7, end_column=13)                  
        #     else:
        #         self.styled_cell(self.ws.cell(7, 1).coordinate, 'Исполнитель: ', 'simple_text_bold3') 
        #         self.ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=6)  
                
        #         self.styled_cell(self.ws.cell(7, 7).coordinate, 'Время начала сушки: ', 'simple_text_bold3') 
        #         self.ws.merge_cells(start_row=7, start_column=7, end_row=7, end_column=13)      
                

        #     self.styled_cell(self.ws.cell(8, 1).coordinate, 'Дата ', 'simple_text_bold3') 
        #     self.styled_cell(self.ws.cell(8, 2).coordinate, 'Смена ', 'simple_text_bold3') 
        #     self.styled_cell(self.ws.cell(8, 3).coordinate, '№ цикла ', 'simple_text_bold3') 
        #     self.styled_cell(self.ws.cell(8, 4).coordinate, '№ б/б ', 'simple_text_bold3') 
        #     self.styled_cell(self.ws.cell(8, 5).coordinate, '№ б/б ', 'simple_text_bold3') 
        #     self.styled_cell(self.ws.cell(8, 6).coordinate, 'Вес  лотка, кг ', 'simple_text_bold3') 
        #     self.styled_cell(self.ws.cell(8, 7).coordinate, 'Влажный вес с  лотком, кг ', 'simple_text_bold3') 
        #     self.styled_cell(self.ws.cell(8, 8).coordinate, 'Вес после первоначальной сушки* с лотком, кг ', 'simple_text_bold3') 
        #     self.styled_cell(self.ws.cell(8, 9).coordinate, 'Вес после доп. часа сушки с лотком, кг ', 'simple_text_bold3') 
        #     self.styled_cell(self.ws.cell(8, 10).coordinate, 'Чистый влажный вес, кг ', 'simple_text_bold3') 
        #     self.styled_cell(self.ws.cell(8, 11).coordinate, 'Чистый сухой вес, кг ', 'simple_text_bold3') 
        #     self.styled_cell(self.ws.cell(8, 12).coordinate, 'Содержание влаги, % ', 'simple_text_bold3') 

   
        #     # дата
        #     self.styled_cell(self.ws.cell(9, 1).coordinate, i, 'simple_text_bold3')
        #     self.ws.merge_cells(start_row=9, start_column=1, end_row=9+len(data_all), end_column=1)
            
        #     # смена
        #     data_all_ns = data_all.filter(type_of_shift=0)

        #     data_all_ds = data_all.filter(type_of_shift=1)

        #     # n/s
        #     self.styled_cell(self.ws.cell(9, 2).coordinate, 'н/с', 'simple_text_bold3')
            
        #     if len(data_all_ns)>0:
        #         self.ws.merge_cells(start_row=9, start_column=2, end_row=8+len(data_all_ns), end_column=2)
                
        #         # стиль для разделителя
        #         for m in range(2, 13):
        #             self.styled_cell(self.ws.cell(9+len(data_all_ns), m).coordinate, '', 'separator')

            
        #     for obj in data_all_ns:
        #         print('type(obj)', type(obj))
        #         # for m in range(3, 13):
        #         self.styled_cell(self.ws.cell(9+coord, 3).coordinate, obj.cycle_number, 'simple_text')
        #         self.styled_cell(self.ws.cell(9+coord, 4).coordinate, obj.bb1, 'simple_text')
        #         self.styled_cell(self.ws.cell(9+coord, 5).coordinate, obj.bb2, 'simple_text')
        #         self.styled_cell(self.ws.cell(9+coord, 6).coordinate, obj.tray_weight, 'simple_text')
        #         self.styled_cell(self.ws.cell(9+coord, 7).coordinate, obj.wet_weight_w_tray, 'simple_text')
        #         self.styled_cell(self.ws.cell(9+coord, 8).coordinate, obj.w_after_init_dry, 'simple_text')
        #         self.styled_cell(self.ws.cell(9+coord, 9).coordinate, obj.w_after_extra_drying, 'simple_text')
        #         self.styled_cell(self.ws.cell(9+coord, 10).coordinate, obj.net_wet_weight, 'simple_text')
        #         self.styled_cell(self.ws.cell(9+coord, 11).coordinate, obj.net_dry_weight, 'simple_text')
        #         self.styled_cell(self.ws.cell(9+coord, 12).coordinate, obj.moisture_contents, 'simple_text')
        #         coord += 1

        #     coord += 1
        #     # d/s
        #     self.styled_cell(self.ws.cell(10+len(data_all_ns), 2).coordinate, 'д/с', 'simple_text_bold3')
        #     if len(data_all_ds)>0:
        #         self.ws.merge_cells(start_row=10+len(data_all_ns), start_column=2, end_row=9+len(data_all_ns)+len(data_all_ds), end_column=2)
                
        #         for obj in data_all_ds:
        #             print('type(obj)', type(obj))
        #             # for m in range(3, 13):
        #             self.styled_cell(self.ws.cell(9+coord, 3).coordinate, obj.cycle_number, 'simple_text')
        #             self.styled_cell(self.ws.cell(9+coord, 4).coordinate, obj.bb1, 'simple_text')
        #             self.styled_cell(self.ws.cell(9+coord, 5).coordinate, obj.bb2, 'simple_text')
        #             self.styled_cell(self.ws.cell(9+coord, 6).coordinate, obj.tray_weight, 'simple_text')
        #             self.styled_cell(self.ws.cell(9+coord, 7).coordinate, obj.wet_weight_w_tray, 'simple_text')
        #             self.styled_cell(self.ws.cell(9+coord, 8).coordinate, obj.w_after_init_dry, 'simple_text')
        #             self.styled_cell(self.ws.cell(9+coord, 9).coordinate, obj.w_after_extra_drying, 'simple_text')
        #             self.styled_cell(self.ws.cell(9+coord, 10).coordinate, obj.net_wet_weight, 'simple_text')
        #             self.styled_cell(self.ws.cell(9+coord, 11).coordinate, obj.net_dry_weight, 'simple_text')
        #             self.styled_cell(self.ws.cell(9+coord, 12).coordinate, obj.moisture_contents, 'simple_text')
        #             coord += 1
        
        # # Удаляем пустую книгу    
        # std=self.wb.get_sheet_by_name('Sheet')
        # self.wb.remove_sheet(std)
   
    def add_footer(self):
        super().add_footer()
