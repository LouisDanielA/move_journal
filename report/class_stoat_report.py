# import datetime
import json
import os
from pathlib import Path
import openpyxl
# from typing import TextIO
# import loguru
from openpyxl.cell.cell import Cell
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.fonts import Font
from openpyxl.workbook.workbook import Workbook

# from report.laboratory.helpers import BASE_DIR


BASE_DIR = Path(__file__).resolve().parent.parent

class stoat_report:
    """ родительский класс для отчетов """
    name: str = 'base'
    template: str = 'empty'
    app: str = 'empty'

    def __init__(self, start, stop=None) -> None:
        if self.template == 'empty':
            self.wb = Workbook()
        else:
            self.wb = openpyxl.load_workbook(filename=self.template)
        
        self.ws = self.wb.active
        f = open(str(BASE_DIR)+'\\'+self.app+'\\'+self.name+'.json')
        self.conf = json.load(f)
        self.start = start
        self.stop = stop

    def styled_cell(self, cell, value, style) -> Cell:
        cur = self.ws[cell]
        cur.value = value
        cur.font = Font(**self.conf['style'][style]['font'])
        cur.alignment = Alignment(**self.conf['style'][style]['alignment'])
        cur.border = Border(
            **dict(map(lambda i: (i, Side(**self.conf['style'][style]['border'][i])), self.conf['style'][style]['border'])))
        cur.fill = PatternFill(**self.conf['style'][style]['fill'])
        if value != '':
            cur.value = value
        return cur

    def build(self):
        self.add_header()
        self.add_table()
        self.add_footer()

        # dict(map(lambda i: (i, self.ws.merge_cells(i)),
        #      self.conf['cells']['merge']['names']))

        #высота столбца
        for i in self.conf['cells']['width']:
            self.ws.column_dimensions[i].width = self.conf['cells']['width'][i]
            

        xlname = self.name+'.xlsx'
        full_path = os.path.join(BASE_DIR, 'media/Unload/UnloadUser/',
                        f'{self.name}_{str(self.start)}.xlsx')
        self.wb.save(full_path)
        # print('xlname',full_path)
        return full_path

    def add_header(self):
        # logger.debug('{} add_header'.format(self.name))
        pass

    def add_footer(self):
        # logger.debug('{} add_footer'.format(self.name))
        # восстановить потом
        # dict(map(lambda i: (i, self.ws.merge_cells(i)),
            #  self.conf['cells']['merge']['names']))
        # --------------
        pass

    def add_table(self):
        # logger.debug('{} add_table'.format(self.name))
        pass


